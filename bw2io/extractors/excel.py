import os
from openpyxl import load_workbook


def get_cell_value_handle_error(cell):
    if cell.data_type == "e":
        # Error type
        return None
    else:
        return cell.value


class ExcelExtractor(object):
    @classmethod
    def extract(cls, filepath):
        assert os.path.exists(filepath), "Can't file file at path {}".format(filepath)
        wb = load_workbook(filepath, data_only=True)
        return [(name, cls.extract_sheet(wb, name)) for name in wb.sheetnames]

    @classmethod
    def extract_sheet(cls, wb, name, strip=True):
        ws = wb[name]
        _ = lambda x: x.strip() if (strip and hasattr(x, "strip")) else x
        return [
            [
                _(get_cell_value_handle_error(cell))
                for colidx, cell in enumerate(row)
            ]
            for rowidx, row in enumerate(ws.rows)
        ]
